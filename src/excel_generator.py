"""Excel spreadsheet generation for planning documents."""

from datetime import datetime, timedelta

import click
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Colors matching the original spreadsheet
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
RED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

# Status types that should NOT show estimation points when no cycle is assigned
INACTIVE_STATUS_TYPES = {"backlog", "triage", "canceled", "cancelled"}
# Status types that SHOULD show estimation points when no cycle is assigned
ACTIVE_STATUS_TYPES = {"started", "completed", "unstarted"}


def generate_week_dates(start_date: datetime, num_weeks: int) -> list:
    """Generate a list of week start dates."""
    return [start_date + timedelta(weeks=i) for i in range(num_weeks)]


def format_name(name: str, first_only: bool = True) -> str:
    """Convert email-style names to proper names.

    Examples (first_only=True):
        'john@company.com' -> 'John'
        'john.doe@company.com' -> 'John'
        'John Doe' -> 'John'
    Examples (first_only=False):
        'john.doe@company.com' -> 'John Doe'
    """
    if not name:
        return name

    # If it looks like an email, extract the name part
    if "@" in name:
        name = name.split("@")[0]
        # Convert dots to spaces and capitalize each word
        name = " ".join(word.capitalize() for word in name.split("."))

    if first_only:
        return name.split()[0] if name else name

    return name


def extract_unique_assignees(issues: list) -> list:
    """Extract unique assignee names from issues, formatted as proper names."""
    raw_names = {
        issue["assignee"]["name"]
        for issue in issues
        if issue.get("assignee") and issue["assignee"].get("name")
    }
    return sorted(format_name(name) for name in raw_names)


def get_week_index(cycle_start: str, start_date: datetime, num_weeks: int) -> int:
    """Calculate which week column index (0-based) a cycle falls into.

    Returns:
        - The week index (0-based) if valid
        - -1 if cycle_start is invalid/empty
        - -2 if cycle is before start_date
        - num_weeks (or higher) if cycle is after the week range
    """
    if not cycle_start:
        return -1
    try:
        cycle_date = datetime.fromisoformat(cycle_start.replace("Z", "+00:00")).replace(tzinfo=None)
        days_diff = (cycle_date - start_date).days
        week_index = days_diff // 7
        if days_diff < 0:
            return -2  # Before start_date
        return week_index  # Could be >= num_weeks if after the range
    except (ValueError, TypeError):
        return -1


def populate_sheet(
    ws,
    team_name: str,
    quarter: str,
    issues: list,
    start_date: datetime,
    num_weeks: int = 13,
    cycle_start: str = None,
):
    """Populate a worksheet with planning data.

    If cycle_start is provided, only issues from cycles up to and including
    that date will have estimates placed in the weekly columns.
    All issues are still shown.
    """

    # Styles
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Get unique assignees from issues
    assignees = extract_unique_assignees(issues)

    # Row 1: Capacity header and week dates
    capacity_header_row = 1
    ws.cell(row=capacity_header_row, column=7, value="Assignee")
    ws.cell(row=capacity_header_row, column=7).font = header_font
    ws.cell(row=capacity_header_row, column=7).fill = YELLOW_FILL

    week_dates = generate_week_dates(start_date, num_weeks)
    for i, date in enumerate(week_dates):
        cell = ws.cell(row=capacity_header_row, column=8 + i, value=date)
        cell.number_format = "m/d"
        cell.font = header_font
        cell.fill = YELLOW_FILL

    # Add "Capacity/week" column after week dates
    capacity_week_col = 8 + num_weeks
    cell = ws.cell(row=capacity_header_row, column=capacity_week_col, value="Capacity/week")
    cell.font = header_font
    cell.fill = YELLOW_FILL

    # Rows 2+: Engineer capacity rows (SUMIF formulas)
    capacity_start_row = 2

    # Calculate header row position (immediately after capacity section with 1 blank row)
    header_row = capacity_start_row + len(assignees) + 1

    # Data rows start right after header
    data_start_row = header_row + 1
    estimated_last_row = data_start_row + len(issues)

    for idx, assignee_name in enumerate(assignees):
        row = capacity_start_row + idx
        ws.cell(row=row, column=7, value=assignee_name)

        for i in range(num_weeks):
            col = 8 + i
            col_letter = get_column_letter(col)
            formula = f'=SUMIF($G${data_start_row}:$G${estimated_last_row},$G{row},{col_letter}${data_start_row}:{col_letter}${estimated_last_row})'
            ws.cell(row=row, column=col, value=formula)

    # Header row
    headers = [
        ("A", "Initiative"),
        ("B", "Projects"),
        ("C", "Issue"),
        ("D", "Estimate (days)"),
        ("E", "Description"),
        ("F", "Linear Ticket"),
        ("G", "Assigned to"),
    ]

    for col_letter, header_text in headers:
        cell = ws[f"{col_letter}{header_row}"]
        cell.value = header_text
        cell.font = header_font
        cell.border = thin_border
        cell.fill = YELLOW_FILL

    # Week date headers in header row (use actual dates, not formulas)
    for i, date in enumerate(week_dates):
        col = 8 + i
        cell = ws.cell(row=header_row, column=col)
        cell.value = date
        cell.number_format = "m/d"
        cell.font = header_font
        cell.border = thin_border
        cell.fill = YELLOW_FILL

    # Group issues by initiative and project
    grouped_issues = {}
    for issue in issues:
        project = issue.get("project") or {}
        project_name = project.get("name", "No Project")
        initiatives = project.get("initiatives", {}).get("nodes", [])
        initiative_name = initiatives[0].get("name") if initiatives else "No Initiative"

        key = (initiative_name, project_name)
        grouped_issues.setdefault(key, []).append(issue)

    # Write issues with gray separator rows between initiatives
    current_row = data_start_row
    last_initiative = None
    last_col = 7 + num_weeks  # Last column for gray fill

    for initiative_name, project_name in sorted(grouped_issues.keys()):
        # Add gray separator row when initiative changes (except for first)
        if last_initiative is not None and initiative_name != last_initiative:
            for col in range(1, last_col + 1):
                ws.cell(row=current_row, column=col).fill = GRAY_FILL
            current_row += 1

        last_initiative = initiative_name

        for issue in grouped_issues[(initiative_name, project_name)]:
            ws.cell(row=current_row, column=1, value=initiative_name)
            ws.cell(row=current_row, column=2, value=project_name)
            ws.cell(row=current_row, column=3, value=issue.get("title", ""))

            issue_cycle = issue.get("cycle") or {}
            estimate = issue.get("estimate")
            if estimate is not None:
                cell = ws.cell(row=current_row, column=4, value=float(estimate))
                cell.fill = GREEN_FILL

            description = issue.get("description") or ""
            ws.cell(row=current_row, column=5, value=description[:500] if len(description) > 500 else description)

            cell = ws.cell(row=current_row, column=6, value=issue.get("url", ""))
            cell.alignment = Alignment(wrap_text=True)

            assignee = issue.get("assignee") or {}
            assignee_name = format_name(assignee.get("name", ""))

            # Get issue status type
            issue_state = issue.get("state") or {}
            status_type = issue_state.get("type", "").lower()

            # Handle missing assignee - show "No assignee" with red background
            if not assignee_name:
                ws.cell(row=current_row, column=7, value="No assignee")
                ws.cell(row=current_row, column=7).fill = RED_FILL
            else:
                ws.cell(row=current_row, column=7, value=assignee_name)

            # Fill weekly capacity based on issue's cycle
            # If cycle_id is set (by-cycles mode), only show issues up to and including this cycle
            # Otherwise show all issues
            issue_cycle_start = issue_cycle.get("startsAt", "")
            if estimate is not None:
                if issue_cycle_start:
                    # Issue has cycle: place at the cycle's week with green fill
                    # Check if this issue's cycle should be shown on this tab
                    should_show = True
                    if cycle_start:
                        # Only show if issue's cycle starts on or before this tab's cycle
                        should_show = issue_cycle_start <= cycle_start

                    if should_show:
                        week_idx = get_week_index(issue_cycle_start, start_date, num_weeks)
                        # Handle out-of-range: clamp to valid range
                        if week_idx < 0:
                            week_idx = 0
                        elif week_idx >= num_weeks:
                            week_idx = num_weeks - 1
                        cell = ws.cell(row=current_row, column=8 + week_idx, value=float(estimate))
                        cell.fill = GREEN_FILL
                elif status_type not in INACTIVE_STATUS_TYPES:
                    # Issue has no cycle but is in active status: use updatedAt date
                    # Show estimate with "(No cycle!)" suffix and yellow fill
                    updated_at = issue.get("updatedAt", "")
                    if updated_at:
                        week_idx = get_week_index(updated_at, start_date, num_weeks)
                        # Clamp to valid range
                        if week_idx < 0:
                            week_idx = 0
                        elif week_idx >= num_weeks:
                            week_idx = num_weeks - 1
                    else:
                        # Fallback to last week if no updatedAt
                        week_idx = num_weeks - 1
                    # Show estimate with "(No cycle!)" indicator
                    cell = ws.cell(row=current_row, column=8 + week_idx, value=f"{int(estimate)} (No cycle!)")
                    cell.fill = YELLOW_FILL
                # else: Issue is in backlog/canceled status with no cycle - don't show estimate anywhere

            current_row += 1

    # Update SUMIF formulas with actual row range
    actual_last_row = current_row - 1
    for idx in range(len(assignees)):
        row = capacity_start_row + idx
        for i in range(num_weeks):
            col = 8 + i
            col_letter = get_column_letter(col)
            formula = f'=SUMIF($G${data_start_row}:$G${actual_last_row},$G{row},{col_letter}${data_start_row}:{col_letter}${actual_last_row})'
            ws.cell(row=row, column=col, value=formula)

    # Column widths (A=Initiative, B=Projects, C=Issue, D=Estimate, E=Description, F=Linear Ticket, G=Assigned to)
    widths = {"A": 30, "B": 35, "C": 50, "D": 15, "E": 50, "F": 40, "G": 15}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for i in range(num_weeks + 1):
        ws.column_dimensions[get_column_letter(8 + i)].width = 8


def create_excel(
    team_name: str,
    quarter: str,
    issues: list,
    output_file: str,
    start_date: datetime,
    num_weeks: int = 13,
):
    """Create a new Excel planning spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = start_date.strftime("%m-%d")

    populate_sheet(ws, team_name, quarter, issues, start_date, num_weeks)

    wb.save(output_file)
    click.echo(f"Excel file saved to: {output_file}")


def refresh_excel(
    team_name: str,
    quarter: str,
    issues: list,
    input_file: str,
    start_date: datetime,
    num_weeks: int = 13,
):
    """Refresh an existing Excel file with latest Linear data.

    For issues WITH assignee in Linear: refresh all data including estimate placement
    For issues WITHOUT assignee in Linear: preserve existing assignee and estimate placement
    """
    wb = load_workbook(input_file)

    # Read existing data from the file
    existing_capacity, existing_assignees, existing_num_weeks = read_existing_capacity_data(wb, start_date)

    # Extend num_weeks to match existing file if it has more columns
    if existing_num_weeks > num_weeks:
        click.echo(f"Extending weeks from {num_weeks} to {existing_num_weeks} to match existing file")
        num_weeks = existing_num_weeks

    old_sheet = wb.active
    old_title = old_sheet.title

    # Remove old sheet and create fresh one
    wb.remove(old_sheet)
    ws = wb.create_sheet(title=old_title, index=0)

    populate_sheet_refresh(
        ws, team_name, quarter, issues, start_date, num_weeks,
        existing_capacity=existing_capacity, existing_assignees=existing_assignees
    )

    wb.save(input_file)
    click.echo(f"Excel file refreshed: {input_file}")


def populate_sheet_refresh(
    ws,
    team_name: str,  # noqa: ARG001 - kept for API consistency
    quarter: str,  # noqa: ARG001 - kept for API consistency
    issues: list,
    start_date: datetime,
    num_weeks: int = 13,
    existing_capacity: dict = None,
    existing_assignees: dict = None,
):
    """Populate a worksheet with refreshed Linear data.

    For issues WITH assignee in Linear: use Linear data for assignee and estimate placement
    For issues WITHOUT assignee in Linear: preserve existing assignee and estimate placement from Excel
    """
    if existing_capacity is None:
        existing_capacity = {}
    if existing_assignees is None:
        existing_assignees = {}

    # Styles
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Get unique assignees from issues and existing Excel data
    assignees = extract_unique_assignees(issues)

    # Add existing assignees from Excel that aren't already in the list
    for existing_assignee in existing_assignees.values():
        if existing_assignee and existing_assignee not in assignees:
            assignees.append(existing_assignee)

    week_dates = generate_week_dates(start_date, num_weeks)

    # Track which columns have Linear data vs manual estimates (will be filled during data processing)
    # Key: week_idx, Value: set of sources ("Linear" or "Estimated")
    column_sources = {i: set() for i in range(num_weeks)}

    # Row layout:
    # Row 1: Data Source row (Linear/Estimation per column)
    # Row 2: Capacity header with week dates
    # Rows 3+: Engineer capacity SUMIF rows
    # Blank separator row
    # Header row with column labels
    # Data rows

    source_indicator_row = 1
    capacity_header_row = 2
    capacity_start_row = 3

    # Write Capacity header row
    ws.cell(row=capacity_header_row, column=8, value="Assignee")
    ws.cell(row=capacity_header_row, column=8).font = header_font
    ws.cell(row=capacity_header_row, column=8).fill = YELLOW_FILL

    for i, date in enumerate(week_dates):
        cell = ws.cell(row=capacity_header_row, column=9 + i, value=date)
        cell.number_format = "m/d"
        cell.font = header_font
        cell.fill = YELLOW_FILL

    # Add "Capacity/week" column after week dates
    capacity_week_col = 9 + num_weeks
    cell = ws.cell(row=capacity_header_row, column=capacity_week_col, value="Capacity/week")
    cell.font = header_font
    cell.fill = YELLOW_FILL

    # Calculate remaining row positions
    # - capacity_start_row + len(assignees) - 1: last capacity row
    # - +1: blank separator row
    # - +1: header row with column labels
    header_row = capacity_start_row + len(assignees) + 1

    # Data rows start right after header
    data_start_row = header_row + 1
    estimated_last_row = data_start_row + len(issues)

    for idx, assignee_name in enumerate(assignees):
        row = capacity_start_row + idx
        ws.cell(row=row, column=8, value=assignee_name)

        for i in range(num_weeks):
            col = 9 + i
            col_letter = get_column_letter(col)
            formula = f'=SUMIF($H${data_start_row}:$H${estimated_last_row},$H{row},{col_letter}${data_start_row}:{col_letter}${estimated_last_row})'
            ws.cell(row=row, column=col, value=formula)

    # Header row - Column A shows data source (Linear vs Estimated)
    headers = [
        ("A", "Linear vs Estimated"),
        ("B", "Initiative"),
        ("C", "Projects"),
        ("D", "Issue"),
        ("E", "Estimate (days)"),
        ("F", "Description"),
        ("G", "Linear Ticket"),
        ("H", "Assigned to"),
    ]

    for col_letter, header_text in headers:
        cell = ws[f"{col_letter}{header_row}"]
        cell.value = header_text
        cell.font = header_font
        cell.border = thin_border
        cell.fill = YELLOW_FILL

    # Week date headers in header row (use actual dates)
    for i, date in enumerate(week_dates):
        col = 9 + i
        cell = ws.cell(row=header_row, column=col)
        cell.value = date
        cell.number_format = "m/d"
        cell.font = header_font
        cell.border = thin_border
        cell.fill = YELLOW_FILL

    # Group issues by initiative and project
    grouped_issues = {}
    for issue in issues:
        project = issue.get("project") or {}
        project_name = project.get("name", "No Project")
        initiatives = project.get("initiatives", {}).get("nodes", [])
        initiative_name = initiatives[0].get("name") if initiatives else "No Initiative"

        key = (initiative_name, project_name)
        grouped_issues.setdefault(key, []).append(issue)

    # Write issues with gray separator rows between initiatives
    current_row = data_start_row
    last_initiative = None
    last_col = 8 + num_weeks  # Last column for gray fill

    for initiative_name, project_name in sorted(grouped_issues.keys()):
        # Add gray separator row when initiative changes (except for first)
        if last_initiative is not None and initiative_name != last_initiative:
            for col in range(1, last_col + 1):  # Start from column A
                ws.cell(row=current_row, column=col).fill = GRAY_FILL
            current_row += 1

        last_initiative = initiative_name

        for issue in grouped_issues[(initiative_name, project_name)]:
            # Determine data source: "Linear" if Linear has both estimate and assignee, "Estimated" otherwise
            linear_assignee = issue.get("assignee") or {}
            linear_has_assignee = bool(linear_assignee.get("name"))
            linear_has_estimate = issue.get("estimate") is not None
            issue_cycle = issue.get("cycle") or {}
            linear_has_cycle = bool(issue_cycle.get("startsAt", ""))

            # "Linear" if assignee and estimate exist in Linear, otherwise "Estimated"
            data_source = "Linear" if (linear_has_assignee and linear_has_estimate) else "Estimated"
            ws.cell(row=current_row, column=1, value=data_source)

            ws.cell(row=current_row, column=2, value=initiative_name)
            ws.cell(row=current_row, column=3, value=project_name)
            ws.cell(row=current_row, column=4, value=issue.get("title", ""))

            estimate = issue.get("estimate")
            if estimate is not None:
                cell = ws.cell(row=current_row, column=5, value=float(estimate))
                cell.fill = GREEN_FILL

            description = issue.get("description") or ""
            ws.cell(row=current_row, column=6, value=description[:500] if len(description) > 500 else description)

            issue_url = issue.get("url", "")
            cell = ws.cell(row=current_row, column=7, value=issue_url)
            cell.alignment = Alignment(wrap_text=True)

            # Get issue status type
            issue_state = issue.get("state") or {}
            status_type = issue_state.get("type", "").lower()

            if linear_has_assignee:
                # Linear has assignee: use Linear data
                assignee_name = format_name(linear_assignee.get("name", ""))
            else:
                # No assignee in Linear: preserve existing Excel assignee
                assignee_name = existing_assignees.get(issue_url, "")

            # Handle missing assignee - show "No assignee" with red background
            if not assignee_name:
                ws.cell(row=current_row, column=8, value="No assignee")
                ws.cell(row=current_row, column=8).fill = RED_FILL
            else:
                ws.cell(row=current_row, column=8, value=assignee_name)

            # Fill weekly capacity
            issue_cycle_start = issue_cycle.get("startsAt", "")

            # Track which week index came from Linear (if any)
            linear_week_idx = None
            has_linear_placement = False

            if linear_has_cycle and estimate is not None:
                # Linear has cycle: calculate the week position
                linear_week_idx = get_week_index(issue_cycle_start, start_date, num_weeks)
                if linear_week_idx < 0:
                    linear_week_idx = 0
                elif linear_week_idx >= num_weeks:
                    linear_week_idx = num_weeks - 1

                # Linear data gets green fill
                cell = ws.cell(row=current_row, column=9 + linear_week_idx, value=float(estimate))
                cell.fill = GREEN_FILL
                column_sources[linear_week_idx].add("Linear")
                has_linear_placement = True
            elif estimate is not None and status_type not in INACTIVE_STATUS_TYPES:
                # No cycle but active status: use updatedAt date with "(No cycle!)" indicator
                updated_at = issue.get("updatedAt", "")
                if updated_at:
                    week_idx = get_week_index(updated_at, start_date, num_weeks)
                    if week_idx < 0:
                        week_idx = 0
                    elif week_idx >= num_weeks:
                        week_idx = num_weeks - 1
                else:
                    week_idx = num_weeks - 1
                cell = ws.cell(row=current_row, column=9 + week_idx, value=f"{int(estimate)} (No cycle!)")
                cell.fill = YELLOW_FILL
                column_sources[week_idx].add("Estimated")
                has_linear_placement = True  # Don't try to restore from existing file

            # Only preserve existing Excel estimate placements if Linear did NOT place an estimate
            # (i.e., Linear doesn't have cycle info for this issue)
            if not has_linear_placement:
                for week_idx in range(num_weeks):
                    existing_val = existing_capacity.get((issue_url, week_idx))
                    # Only preserve non-zero values (skip 0 or near-zero values)
                    if existing_val is not None and existing_val > 0:
                        # Manual/estimated data gets yellow fill
                        cell = ws.cell(row=current_row, column=9 + week_idx, value=existing_val)
                        cell.fill = YELLOW_FILL
                        column_sources[week_idx].add("Estimated")

            current_row += 1

    # Update SUMIF formulas with actual row range
    actual_last_row = current_row - 1
    for idx in range(len(assignees)):
        row = capacity_start_row + idx
        for i in range(num_weeks):
            col = 9 + i
            col_letter = get_column_letter(col)
            formula = f'=SUMIF($H${data_start_row}:$H${actual_last_row},$H{row},{col_letter}${data_start_row}:{col_letter}${actual_last_row})'
            ws.cell(row=row, column=col, value=formula)

    # Add a row showing "Linear" or "Estimation" for each weekly column (just above the main header)
    ws.cell(row=source_indicator_row, column=8, value="Data Source")
    ws.cell(row=source_indicator_row, column=8).font = header_font

    for week_idx in range(num_weeks):
        sources = column_sources.get(week_idx, set())
        # If Linear data exists in this column, show "Linear" (Linear takes precedence)
        # Only show "Estimation" if there's no Linear data in this column
        if "Linear" in sources:
            source_label = "Linear"
        elif "Estimated" in sources:
            source_label = "Estimation"
        else:
            source_label = ""

        cell = ws.cell(row=source_indicator_row, column=9 + week_idx, value=source_label)
        cell.font = header_font

    # Column widths (A=Linear vs Estimated, G=Linear Ticket fixed width, H=Assigned to)
    widths = {"A": 18, "B": 30, "C": 35, "D": 50, "E": 15, "F": 50, "G": 40, "H": 15}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for i in range(num_weeks + 1):
        ws.column_dimensions[get_column_letter(9 + i)].width = 8


def read_existing_capacity_data(wb, start_date: datetime) -> tuple:
    """Read existing capacity data and assignees from the first sheet of the workbook.

    Returns a tuple of:
    - capacity_data: dict mapping (issue_url, week_index) -> estimate value
    - assignee_data: dict mapping issue_url -> assignee name
    - max_week_idx: the maximum week index found in the existing file (for extending num_weeks)
    """
    if not wb.worksheets:
        return {}, {}, 0

    ws = wb.worksheets[0]
    capacity_data = {}
    assignee_data = {}
    max_week_idx = 0

    # Find header row by looking for "Linear Ticket" in column G
    # Also check column F for older file formats without "Linear vs Estimated" column
    header_row = None
    linear_ticket_col = None
    for row in range(1, min(50, ws.max_row + 1)):
        # Check column G first (new format with "Linear vs Estimated" in column A)
        cell_val_g = ws.cell(row=row, column=7).value
        if cell_val_g and "Linear Ticket" in str(cell_val_g):
            header_row = row
            linear_ticket_col = 7
            break
        # Check column F (old format without "Linear vs Estimated" column)
        cell_val_f = ws.cell(row=row, column=6).value
        if cell_val_f and "Linear Ticket" in str(cell_val_f):
            header_row = row
            linear_ticket_col = 6
            break

    if not header_row:
        return {}, {}, 0

    # Determine column offsets based on file format
    # New format: Linear Ticket in col G (7), Assigned to in col H (8), weeks start at col I (9)
    # Old format: Linear Ticket in col F (6), Assigned to in col G (7), weeks start at col H (8)
    assignee_col = linear_ticket_col + 1
    week_start_col = linear_ticket_col + 2

    # Find the capacity header row by looking for "Capacity" label
    # This row has all the week dates and is more reliable for parsing
    capacity_header_row = None
    for row in range(1, header_row):
        # Check for "Capacity" in column H (8) or nearby columns
        for check_col in range(6, 10):
            cell_val = ws.cell(row=row, column=check_col).value
            if cell_val and str(cell_val).strip() == "Capacity":
                capacity_header_row = row
                break
        if capacity_header_row:
            break

    # Use capacity header row for date parsing if found, otherwise use data header row
    date_header_row = capacity_header_row if capacity_header_row else header_row

    # Find week columns by parsing dates in the header row
    week_col_map = {}  # week_index -> column
    for col in range(week_start_col, ws.max_column + 1):
        cell_val = ws.cell(row=date_header_row, column=col).value
        if cell_val:
            if isinstance(cell_val, datetime):
                week_date = cell_val
            elif isinstance(cell_val, str):
                # Try to parse M/D format
                try:
                    # First try with start_date year
                    week_date = datetime.strptime(f"{cell_val}/{start_date.year}", "%m/%d/%Y")
                    # If the parsed date is before start_date, it's likely next year
                    if week_date < start_date:
                        week_date = datetime.strptime(f"{cell_val}/{start_date.year + 1}", "%m/%d/%Y")
                except ValueError:
                    continue
            else:
                continue

            # Calculate week index
            days_diff = (week_date - start_date).days
            week_idx = days_diff // 7
            if week_idx >= 0:
                week_col_map[week_idx] = col
                if week_idx > max_week_idx:
                    max_week_idx = week_idx

    # Read data rows (start after header row)
    for row in range(header_row + 1, ws.max_row + 1):
        url = ws.cell(row=row, column=linear_ticket_col).value  # Linear Ticket URL
        if not url:
            continue

        # Read assignee
        assignee_val = ws.cell(row=row, column=assignee_col).value
        if assignee_val:
            assignee_data[url] = str(assignee_val)

        # Read capacity values for each week
        for week_idx, col in week_col_map.items():
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is not None and cell_val != "":
                try:
                    val = float(cell_val)
                    # Only store non-zero values
                    if val > 0:
                        capacity_data[(url, week_idx)] = val
                except (ValueError, TypeError):
                    pass

    return capacity_data, assignee_data, max_week_idx + 1  # +1 because week_idx is 0-based
